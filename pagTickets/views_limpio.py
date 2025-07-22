# Importa funciones necesarias de Django
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import RegistroQR

# Importar librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# Función principal que muestra la página de inicio con el escáner QR
def index(request):
    # Obtiene los últimos 10 registros de la base de datos, ordenados por fecha (más reciente primero)
    registros = RegistroQR.objects.order_by('-fecha_registro')[:10]
    # Renderiza la página HTML y le pasa los registros como contexto
    return render(request, 'index.html', {'registros': registros})

# Función que guarda un nuevo código QR en la base de datos
@csrf_exempt
def registrar_qr(request):
    if request.method == 'POST':
        try:
            # Convierte los datos JSON recibidos a un diccionario de Python
            data = json.loads(request.body)
            # Extrae el código QR de los datos recibidos
            codigo = data.get('codigo')
            usuario = data.get('usuario', 'Usuario Web')
            ubicacion = data.get('ubicacion', 'Escáner Web')
            notas = data.get('notas', '')
            
            # Crea un nuevo registro en la base de datos con el código QR
            registro = RegistroQR.objects.create(
                codigo=codigo,
                usuario=usuario,
                ubicacion=ubicacion,
                notas=notas
            )
            
            # Devuelve una respuesta JSON exitosa con el código registrado
            return JsonResponse({
                'success': True, 
                'codigo': registro.codigo,
                'fecha': registro.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
            })
        except Exception as e:
            # Si hay algún error, devuelve una respuesta JSON con el error
            return JsonResponse({'success': False, 'error': str(e)})
    
    # Si no es una petición POST, devuelve un error
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# Función para obtener los últimos registros (para actualizar la lista en tiempo real)
def ultimos_registros(request):
    try:
        # Obtiene los últimos 20 registros
        registros = RegistroQR.objects.order_by('-fecha_registro')[:20]
        
        # Convierte los registros a formato JSON
        registros_data = []
        for registro in registros:
            registros_data.append({
                'codigo': registro.codigo,
                'fecha_registro': registro.fecha_registro.strftime('%Y-%m-%d %H:%M:%S'),
                'usuario': registro.usuario,
                'ubicacion': registro.ubicacion,
                'notas': registro.notas
            })
        
        return JsonResponse({'success': True, 'registros': registros_data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Función para exportar registros a Excel
def exportar_excel(request):
    try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros QR"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados
        headers = ['Código QR', 'Fecha y Hora', 'Usuario', 'Ubicación', 'Notas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Obtener todos los registros
        registros = RegistroQR.objects.order_by('-fecha_registro')
        
        # Agregar datos
        for row, registro in enumerate(registros, 2):
            ws.cell(row=row, column=1, value=registro.codigo).border = border
            ws.cell(row=row, column=2, value=registro.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')).border = border
            ws.cell(row=row, column=3, value=registro.usuario).border = border
            ws.cell(row=row, column=4, value=registro.ubicacion).border = border
            ws.cell(row=row, column=5, value=registro.notas).border = border
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="registros_qr_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Importa funciones necesarias de Django
from django.shortcuts import render  # Para mostrar páginas HTML
from django.http import JsonResponse, HttpResponse  # Para enviar respuestas en formato JSON y HTTP
from django.views.decorators.csrf import csrf_exempt  # Para permitir peticiones POST sin token CSRF
import json  # Para trabajar con datos JSON
from .models import RegistroQR  # Importa el modelo desde models.py

# Importar librerías para Excel
from openpyxl import Workbook  # Para crear archivos Excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Para estilos de Excel
from openpyxl.utils import get_column_letter  # Para obtener letras de columnas
import datetime  # Para trabajar con fechas

# Función principal que muestra la página de inicio con los códigos QR
def index(request):
    # Obtiene los últimos 10 registros de la base de datos, ordenados por fecha (más reciente primero)
    registros = RegistroQR.objects.order_by('-fecha_registro')[:10]
    # Renderiza la página HTML y le pasa los registros como contexto
    return render(request, 'index.html', {'registros': registros})

# Decorador que permite recibir peticiones POST sin validación CSRF
@csrf_exempt
# Función que guarda un nuevo código QR en la base de datos
def registrar_qr(request):
    # Verifica si la petición es de tipo POST (envío de datos)
    if request.method == 'POST':
        try:
            # Convierte los datos JSON recibidos a un diccionario de Python
            data = json.loads(request.body)
            # Extrae el código QR de los datos recibidos
            codigo = data.get('codigo')
            usuario = data.get('usuario', '')
            ubicacion = data.get('ubicacion', '')
            notas = data.get('notas', '')
            
            # Crea un nuevo registro en la base de datos con el código QR
            registro = RegistroQR.objects.create(
                codigo=codigo,
                usuario=usuario,
                ubicacion=ubicacion,
                notas=notas
            )
            
            # Devuelve una respuesta JSON exitosa con el código registrado
            return JsonResponse({
                'success': True, 
                'codigo': registro.codigo,
                'fecha': registro.fecha_registro.strftime('%Y-%m-%d %H:%M:%S'),
                'mensaje': 'Código QR registrado exitosamente'
            })
        except Exception as e:
            # Si hay algún error, devuelve una respuesta JSON con el error
            return JsonResponse({'success': False, 'message': str(e)})
    # Si no es una petición POST, devuelve un error
    return JsonResponse({'success': False, 'message': 'Método no permitido'})

# Función que devuelve los últimos códigos QR registrados en formato JSON
def ultimos_registros(request):
    # Obtiene los últimos 10 registros ordenados por fecha
    registros = RegistroQR.objects.order_by('-fecha_registro')[:10]
    # Convierte cada registro a un diccionario con código y fecha formateada
    data = [
        {
            'codigo': r.codigo, 
            'fecha': r.fecha_registro.strftime('%Y-%m-%d %H:%M:%S'),
            'usuario': r.usuario,
            'ubicacion': r.ubicacion,
            'notas': r.notas
        }
        for r in registros  # Recorre cada registro y crea un diccionario
    ]
    # Devuelve la respuesta JSON con la lista de registros
    return JsonResponse({'registros': data})

# Función para exportar registros QR a Excel con formato profesional
def exportar_excel(request):
    # Obtener todos los registros de la base de datos
    registros = RegistroQR.objects.order_by('-fecha_registro')
    
    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Registros QR"
    
    # Configurar estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='dc2626', end_color='dc2626', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Código QR', 'Fecha y Hora', 'Usuario', 'Ubicación', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row_num, registro in enumerate(registros, 2):
        worksheet.cell(row=row_num, column=1, value=registro.codigo)
        worksheet.cell(row=row_num, column=2, value=registro.fecha_registro.strftime('%d/%m/%Y %H:%M:%S'))
        worksheet.cell(row=row_num, column=3, value=registro.usuario)
        worksheet.cell(row=row_num, column=4, value=registro.ubicacion)
        worksheet.cell(row=row_num, column=5, value=registro.notas)
        
        # Aplicar formato a las celdas de datos
        for col in range(1, 6):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    column_widths = [20, 20, 15, 25, 40]
    for col, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    
    # Agregar información adicional
    last_row = len(registros) + 3
    
    # Celda con total de registros
    total_cell = worksheet.cell(row=last_row, column=1)
    total_cell.value = f"Total de registros: {len(registros)}"
    worksheet.merge_cells(f'A{last_row}:B{last_row}')
    total_cell.font = Font(name='Arial', size=11, bold=True)
    total_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_cell.border = thin_border
    
    # Celda con fecha de exportación
    export_cell = worksheet.cell(row=last_row, column=3)
    export_cell.value = f"Exportado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    worksheet.merge_cells(f'C{last_row}:E{last_row}')
    export_cell.font = Font(name='Arial', size=10, italic=True)
    export_cell.alignment = Alignment(horizontal='right', vertical='center')
    export_cell.border = thin_border
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha actual
    filename = f'registros_qr_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Guardar el workbook en la respuesta
    workbook.save(response)
    
    return response

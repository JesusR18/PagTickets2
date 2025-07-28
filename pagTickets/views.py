# Importa funciones necesarias de Django
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
# Imports de auth eliminados - ya no se usan
from django.utils import timezone
import json
from .models import RegistroQR

# Importar librerías para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# Función helper para formatear fechas con zona horaria local
def format_local_datetime(dt):
    """
    Convierte un datetime UTC a la zona horaria local configurada en Django
    y lo formatea como string
    """
    if dt:
        # Convertir a zona horaria local
        local_dt = timezone.localtime(dt)
        # Formatear con zona horaria local
        return local_dt.strftime('%Y-%m-%d %H:%M:%S')
    return ''

# Función helper para obtener código según ubicación
def obtener_codigo_ubicacion(ubicacion):
    """
    Determina el código específico según la ubicación del activo
    """
    if not ubicacion:
        return "ARC"
    
    # Convertir a minúsculas para comparación case-insensitive
    ubicacion_lower = ubicacion.lower().strip()
    
    # Administración
    if ubicacion_lower in ["administracion", "1er piso administracion"]:
        return "ADMON"
    
    # Almacén
    elif ubicacion_lower in ["almacen", "1er piso almacen"]:
        return "ALM"
    
    # Crédito y cobranza
    elif ubicacion_lower in ["credito y cobranza", "cuentas por pagar", "1er piso cuentas por pagar"]:
        return "CRED"
    
    # Dirección
    elif ubicacion_lower in ["direccion", "2do piso direccion"]:
        return "DIR"
    
    # Gerencia
    elif ubicacion_lower in ["gerencia", "gerencia general", "2do piso gerencia general"]:
        return "GER"
    
    # Gerencia de ventas
    elif ubicacion_lower in ["gerencia de ventas", "1er piso gerencia de ventas"]:
        return "GV"
    
    # Proyectos y Marketing
    elif ubicacion_lower in ["proyectos", "1er piso proyectos", "marketing", "1er piso marketing"]:
        return "PROY"
    
    # Monitoreo, SITE, Video Wall
    elif ubicacion_lower in ["monitoreo", "2do piso monitoreo", "site", "2do piso site", "video wall", "2do piso video wall"]:
        return "MON"
    
    # Recursos Humanos
    elif ubicacion_lower in ["1er piso r.h"]:
        return "R.H."
    
    # Sala de juntas
    elif ubicacion_lower in ["sala de juntas", "2do piso sala juntas"]:
        return "SJ"
    
    # Ventas
    elif ubicacion_lower in ["ventas", "1er piso ventas"]:
        return "VEN 1"
    
    # Archivo (default)
    else:
        return "ARC"

# Vista de healthcheck para Railway
def health_check(request):
    """Vista mejorada para verificación de salud de Railway"""
    try:
        import django
        from django.db import connection
        from django.conf import settings
        import json
        
        health_data = {
            "status": "ok",
            "django_version": django.get_version(),
            "debug": settings.DEBUG,
            "database": "disconnected",
            "timestamp": timezone.now().isoformat()
        }
        
        # Verificación de base de datos (más robusta)
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                row = cursor.fetchone()
                if row and row[0] == 1:
                    health_data["database"] = "connected"
        except Exception as db_error:
            health_data["database"] = f"error: {str(db_error)}"
        
        # Si es una petición simple, devolver solo OK
        if request.GET.get('simple') == 'true':
            return HttpResponse("OK", content_type="text/plain", status=200)
        
        # Devolver información detallada en JSON
        return HttpResponse(
            json.dumps(health_data, indent=2), 
            content_type="application/json", 
            status=200
        )
        
    except Exception as e:
        # En caso de error, devolver status 500
        error_response = {
            "status": "error",
            "message": str(e),
            "timestamp": timezone.now().isoformat()
        }
        return HttpResponse(
            json.dumps(error_response, indent=2), 
            content_type="application/json", 
            status=500
        )

# Vista ultra-simple para ping/healthcheck
def simple_ping(request):
    """Vista ultra-simple que siempre responde OK"""
    return HttpResponse("PONG", content_type="text/plain", status=200)

# Función principal que muestra la página de inicio con el escáner QR
def index(request):
    # Obtiene los últimos activos escaneados desde RegistroQR
    registros_qr = RegistroQR.objects.order_by('-fecha_registro')[:50]
    
    # Extraer información de activos desde los códigos QR
    activos_escaneados = []
    for registro in registros_qr:
        try:
            # Intentar parsear el código QR como JSON
            qr_data = json.loads(registro.codigo)
            activo_info = {
                'codigo': qr_data.get('codigo', registro.codigo),
                'nombre': qr_data.get('nombre', 'Activo sin nombre'),
                'ubicacion': qr_data.get('ubicacion', 'Sin ubicación'),
                'marca': qr_data.get('marca', 'Sin marca'),
                'modelo': qr_data.get('modelo', 'Sin modelo'),
                'no_serie': qr_data.get('no_serie', 'Sin número de serie'),
                'fecha_registro': format_local_datetime(registro.fecha_registro)
            }
        except (json.JSONDecodeError, AttributeError):
            # Si no es JSON válido, usar información básica
            activo_info = {
                'codigo': registro.codigo,
                'nombre': registro.codigo,
                'ubicacion': registro.ubicacion or 'Sin ubicación',
                'marca': 'Sin marca',
                'modelo': 'Sin modelo',
                'no_serie': 'Sin número de serie',
                'fecha_registro': format_local_datetime(registro.fecha_registro)
            }
        activos_escaneados.append(activo_info)
    
    # Renderiza la página HTML y le pasa los activos como contexto
    return render(request, 'index.html', {'activos_escaneados': activos_escaneados})

# Función para verificar si un activo ya existe con las mismas características
def verificar_activo_existente(activo_info):
    """
    Verifica si ya existe un activo registrado con las mismas características principales.
    Compara: nombre, ubicación, marca, modelo y número de serie.
    """
    try:
        # Obtener todos los registros para comparar
        registros = RegistroQR.objects.all()
        
        for registro in registros:
            try:
                # Extraer información del registro existente
                info_existente = extraer_informacion_qr(registro.codigo)
                
                # Comparar características principales
                nombre_match = activo_info.get('nombre', '').strip().lower() == info_existente.get('nombre', '').strip().lower()
                ubicacion_match = activo_info.get('ubicacion', '').strip().lower() == info_existente.get('ubicacion', '').strip().lower()
                marca_match = activo_info.get('marca', '').strip().lower() == info_existente.get('marca', '').strip().lower()
                modelo_match = activo_info.get('modelo', '').strip().lower() == info_existente.get('modelo', '').strip().lower()
                
                # Si el número de serie existe y no está vacío, también compararlo
                no_serie_nuevo = activo_info.get('no_serie', '').strip()
                no_serie_existente = info_existente.get('no_serie', '').strip()
                no_serie_match = True  # Por defecto asumimos que coincide
                
                if no_serie_nuevo and no_serie_existente and no_serie_nuevo.lower() != 'sin número de serie':
                    no_serie_match = no_serie_nuevo.lower() == no_serie_existente.lower()
                
                # Si todas las características principales coinciden, es el mismo activo
                if nombre_match and ubicacion_match and marca_match and modelo_match and no_serie_match:
                    return registro
                    
            except Exception:
                continue  # Si hay error parseando un registro, continuar con el siguiente
                
        return None  # No se encontró ningún activo duplicado
        
    except Exception:
        return None  # En caso de error, no bloquear el registro

# Función que guarda un nuevo código QR en la base de datos
@csrf_exempt
def registrar_qr(request):
    if request.method == 'POST':
        try:
            # Convierte los datos JSON recibidos a un diccionario de Python
            data = json.loads(request.body)
            # Extrae el código QR de los datos recibidos
            codigo_qr = data.get('codigo_qr') or data.get('codigo')
            
            # Validar que el código QR no esté vacío
            if not codigo_qr or codigo_qr.strip() == '':
                return JsonResponse({
                    'success': False,
                    'error': 'Código QR vacío o inválido'
                })
            
            usuario = data.get('usuario', 'Usuario Web')
            ubicacion_scan = data.get('ubicacion', 'Escáner Web')
            
            # Intenta parsear el QR como JSON para extraer información del activo
            activo_info = extraer_informacion_qr(codigo_qr)
            
            # Verificar si ya existe un activo con las mismas características
            registro_existente = verificar_activo_existente(activo_info)
            
            if registro_existente:
                # Si ya existe, devolver información de que está registrado
                activo_existente = extraer_informacion_qr(registro_existente.codigo)
                activo_existente['id'] = registro_existente.id
                activo_existente['fecha_registro'] = format_local_datetime(registro_existente.fecha_registro)
                return JsonResponse({
                    'success': True,
                    'already_registered': True,
                    'activo': activo_existente,
                    'mensaje': f'El activo "{activo_existente["nombre"]}" ya está registrado con estas características'
                })
            
            # Si no existe, crear nuevo registro
            nuevo_registro = RegistroQR.objects.create(
                codigo=codigo_qr,
                usuario=usuario,
                ubicacion=ubicacion_scan,
                notas=f"Activo registrado: {activo_info['nombre']}"
            )
            
            # Agregar ID y fecha al activo_info
            activo_info['id'] = nuevo_registro.id
            activo_info['fecha_registro'] = format_local_datetime(nuevo_registro.fecha_registro)
            
            return JsonResponse({
                'success': True,
                'already_registered': False,
                'activo': activo_info,
                'mensaje': f'Activo "{activo_info["nombre"]}" registrado correctamente'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# Función para eliminar un activo
@csrf_exempt
def eliminar_activo(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            activo_id = data.get('id')
            
            if not activo_id:
                return JsonResponse({'success': False, 'error': 'ID de activo requerido'})
            
            # Buscar y eliminar el registro
            registro = RegistroQR.objects.get(id=activo_id)
            activo_info = extraer_informacion_qr(registro.codigo)
            nombre_activo = activo_info['nombre']
            
            registro.delete()
            
            return JsonResponse({
                'success': True,
                'mensaje': f'Activo "{nombre_activo}" eliminado correctamente'
            })
            
        except RegistroQR.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Activo no encontrado'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

# Función auxiliar para extraer información del QR
def extraer_informacion_qr(codigo_qr):
    """
    Extrae información del código QR soportando múltiples formatos:
    - JSON: {"codigo": "ACT001", "nombre": "Laptop", ...}
    - Formato de texto estructurado: "Activo: Escritorio en L Ubicación: 1er piso R.H. Marca: Techni mobili Modelo: Havano N. Serie: -."
    - Texto simple separado por |: ACT001|Laptop Dell|Oficina 1|Dell|Latitude|SN123456
    - Texto simple: cualquier código
    """
    try:
        # Intento 1: Parsear como JSON
        if codigo_qr.strip().startswith('{') and codigo_qr.strip().endswith('}'):
            qr_data = json.loads(codigo_qr)
            ubicacion = qr_data.get('ubicacion', qr_data.get('location', 'Sin ubicación'))
            return {
                'codigo': qr_data.get('codigo', qr_data.get('code', codigo_qr)),
                'nombre': qr_data.get('nombre', qr_data.get('activo', qr_data.get('asset', 'Activo sin nombre'))),
                'ubicacion': ubicacion,
                'marca': qr_data.get('marca', qr_data.get('brand', 'Sin marca')),
                'modelo': qr_data.get('modelo', qr_data.get('model', 'Sin modelo')),
                'no_serie': qr_data.get('no_serie', qr_data.get('serial', qr_data.get('serie', 'Sin número de serie'))),
                'codigo_ubicacion': obtener_codigo_ubicacion(ubicacion)
            }
        
        # Intento 2: Formato de texto estructurado (el formato que mencionas)
        elif any(palabra in codigo_qr.lower() for palabra in ['activo:', 'ubicación:', 'marca:', 'modelo:', 'serie:']):
            return parsear_texto_estructurado(codigo_qr)
        
        # Intento 3: Parsear como texto separado por |
        elif '|' in codigo_qr:
            partes = codigo_qr.split('|')
            ubicacion = partes[2].strip() if len(partes) > 2 else 'Sin ubicación'
            return {
                'codigo': partes[0].strip() if len(partes) > 0 else codigo_qr,
                'nombre': partes[1].strip() if len(partes) > 1 else 'Activo sin nombre',
                'ubicacion': ubicacion,
                'marca': partes[3].strip() if len(partes) > 3 else 'Sin marca',
                'modelo': partes[4].strip() if len(partes) > 4 else 'Sin modelo',
                'no_serie': partes[5].strip() if len(partes) > 5 else 'Sin número de serie',
                'codigo_ubicacion': obtener_codigo_ubicacion(ubicacion)
            }
        
        # Intento 4: Buscar patrones conocidos en el texto
        else:
            # Buscar patrones como "ACT001 - Laptop Dell" o similares
            nombre_detectado = codigo_qr
            codigo_detectado = codigo_qr
            
            # Si contiene guión, probablemente sea código - nombre
            if ' - ' in codigo_qr:
                partes_guion = codigo_qr.split(' - ', 1)
                codigo_detectado = partes_guion[0].strip()
                nombre_detectado = partes_guion[1].strip()
            
            return {
                'codigo': codigo_detectado,
                'nombre': nombre_detectado,
                'ubicacion': 'Sin ubicación',
                'marca': 'Sin marca',
                'modelo': 'Sin modelo',
                'no_serie': 'Sin número de serie',
                'codigo_ubicacion': obtener_codigo_ubicacion('Sin ubicación')
            }
            
    except Exception:
        # Si todo falla, usar el código QR tal como está
        return {
            'codigo': codigo_qr,
            'nombre': codigo_qr,
            'ubicacion': 'Sin ubicación',
            'marca': 'Sin marca',
            'modelo': 'Sin modelo',
            'no_serie': 'Sin número de serie',
            'codigo_ubicacion': obtener_codigo_ubicacion('Sin ubicación')
        }

def parsear_texto_estructurado(texto_qr):
    """
    Parsea texto estructurado como:
    "Activo: Escritorio en L Ubicación: 1er piso R.H. Marca: Techni mobili Modelo: Havano N. Serie: -."
    """
    # Inicializar valores por defecto
    resultado = {
        'codigo': texto_qr[:20] + '...' if len(texto_qr) > 20 else texto_qr,  # Usar parte del texto como código
        'nombre': 'Activo sin nombre',
        'ubicacion': 'Sin ubicación',
        'marca': 'Sin marca',
        'modelo': 'Sin modelo',
        'no_serie': 'Sin número de serie'
    }
    
    try:
        # Convertir texto a minúsculas para búsqueda, pero mantener original para extraer valores
        texto_busqueda = texto_qr.lower()
        
        # Patrones de búsqueda (en orden de prioridad)
        patrones = {
            'nombre': ['activo:', 'asset:', 'equipo:', 'item:'],
            'ubicacion': ['ubicación:', 'ubicacion:', 'location:', 'lugar:'],
            'marca': ['marca:', 'brand:', 'fabricante:'],
            'modelo': ['modelo:', 'model:', 'tipo:'],
            'no_serie': ['n. serie:', 'serie:', 'serial:', 'número de serie:', 'numero de serie:', 'sn:']
        }
        
        for campo, palabras_clave in patrones.items():
            for palabra_clave in palabras_clave:
                if palabra_clave in texto_busqueda:
                    # Encontrar la posición de la palabra clave
                    inicio = texto_busqueda.find(palabra_clave)
                    if inicio != -1:
                        # Buscar el inicio del valor (después de la palabra clave)
                        inicio_valor = inicio + len(palabra_clave)
                        
                        # Buscar el final del valor (hasta la siguiente palabra clave o final del texto)
                        fin_valor = len(texto_qr)
                        
                        # Buscar la siguiente palabra clave para determinar el fin
                        for otros_campos, otras_palabras in patrones.items():
                            if otros_campos != campo:
                                for otra_palabra in otras_palabras:
                                    pos_siguiente = texto_busqueda.find(otra_palabra, inicio_valor)
                                    if pos_siguiente != -1 and pos_siguiente < fin_valor:
                                        fin_valor = pos_siguiente
                        
                        # Extraer el valor y limpiarlo
                        valor = texto_qr[inicio_valor:fin_valor].strip()
                        
                        # Limpiar caracteres innecesarios al final
                        valor = valor.rstrip('.,;:').strip()
                        
                        if valor:
                            resultado[campo] = valor
                        break
        
        # Si no se encontró nombre, usar la primera parte como nombre
        if resultado['nombre'] == 'Activo sin nombre' and resultado['codigo']:
            resultado['nombre'] = resultado['codigo']
        
        # Generar un código más limpio si es posible
        if resultado['nombre'] != 'Activo sin nombre':
            # Tomar las primeras palabras del nombre como código
            palabras_nombre = resultado['nombre'].split()[:3]
            codigo_generado = ''.join(palabra[:3].upper() for palabra in palabras_nombre if palabra.isalpha())
            if len(codigo_generado) >= 3:
                resultado['codigo'] = codigo_generado
        
        # Agregar código de ubicación
        resultado['codigo_ubicacion'] = obtener_codigo_ubicacion(resultado['ubicacion'])
        
        return resultado
        
    except Exception:
        # Si hay error en el parsing, devolver el texto original como nombre
        return {
            'codigo': texto_qr[:20] + '...' if len(texto_qr) > 20 else texto_qr,
            'nombre': texto_qr,
            'ubicacion': 'Sin ubicación',
            'marca': 'Sin marca',
            'modelo': 'Sin modelo',
            'no_serie': 'Sin número de serie',
            'codigo_ubicacion': obtener_codigo_ubicacion('Sin ubicación')
        }

# Función para obtener los activos escaneados (para actualizar la tabla en tiempo real)
def obtener_activos_escaneados(request):
    try:
        # Obtiene todos los registros QR
        registros = RegistroQR.objects.order_by('-fecha_registro')
        
        # Convierte los registros a formato de activos usando la función de extracción mejorada
        activos_data = []
        for registro in registros:
            activo_info = extraer_informacion_qr(registro.codigo)
            # Agregar ID del registro y fecha de registro
            activo_info['id'] = registro.id
            activo_info['fecha_registro'] = format_local_datetime(registro.fecha_registro)
            activos_data.append(activo_info)
        
        return JsonResponse({'activos': activos_data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Función para exportar activos escaneados a Excel
def exportar_activos_excel(request):
    try:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Activos Escaneados"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados
        headers = ['Código QR', 'Activo', 'Ubicación', 'Marca', 'Modelo', 'No. de Serie', 'Fecha de Registro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Obtener todos los registros y convertirlos a activos
        registros = RegistroQR.objects.order_by('-fecha_registro')
        
        # Agregar datos
        row = 2
        for registro in registros:
            # Usar la función de extracción mejorada
            activo_info = extraer_informacion_qr(registro.codigo)
            
            # Combinar código QR con código de ubicación
            codigo_completo = activo_info['codigo']
            if activo_info.get('codigo_ubicacion'):
                codigo_completo += f" ({activo_info['codigo_ubicacion']})"
            
            ws.cell(row=row, column=1, value=codigo_completo).border = border
            ws.cell(row=row, column=2, value=activo_info['nombre']).border = border
            ws.cell(row=row, column=3, value=activo_info['ubicacion']).border = border
            ws.cell(row=row, column=4, value=activo_info['marca']).border = border
            ws.cell(row=row, column=5, value=activo_info['modelo']).border = border
            ws.cell(row=row, column=6, value=activo_info['no_serie']).border = border
            ws.cell(row=row, column=7, value=format_local_datetime(registro.fecha_registro)).border = border
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="activos_escaneados_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# Vista para obtener los últimos registros (API JSON)
def ultimos_registros(request):
    """Vista que devuelve los últimos registros QR en formato JSON"""
    try:
        registros = RegistroQR.objects.order_by('-fecha_registro')[:10]
        datos = []
        
        for registro in registros:
            try:
                # Intentar parsear como JSON
                qr_data = json.loads(registro.codigo)
                datos.append({
                    'id': registro.id,
                    'codigo': qr_data.get('codigo', registro.codigo),
                    'nombre': qr_data.get('nombre', 'Sin nombre'),
                    'ubicacion': qr_data.get('ubicacion', 'Sin ubicación'),
                    'marca': qr_data.get('marca', 'Sin marca'),
                    'modelo': qr_data.get('modelo', 'Sin modelo'),
                    'no_serie': qr_data.get('no_serie', 'Sin número de serie'),
                    'fecha_registro': format_local_datetime(registro.fecha_registro)
                })
            except (json.JSONDecodeError, AttributeError):
                # Si no es JSON, usar el código tal como está
                datos.append({
                    'id': registro.id,
                    'codigo': registro.codigo,
                    'nombre': registro.codigo,
                    'ubicacion': 'Sin ubicación',
                    'marca': 'Sin marca',
                    'modelo': 'Sin modelo',
                    'no_serie': 'Sin número de serie',
                    'fecha_registro': format_local_datetime(registro.fecha_registro)
                })
        
        return JsonResponse({'registros': datos})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Vista para exportar registros QR a Excel
@csrf_exempt
def eliminar_todos_activos(request):
    """Vista para eliminar todos los activos registrados"""
    if request.method == 'POST':
        try:
            # Contar activos antes de eliminar
            total_qr_registros = RegistroQR.objects.count()
            
            # Eliminar todos los registros QR de pagTickets
            RegistroQR.objects.all().delete()
            
            # Eliminar todos los registros QR de qrweb (si existen)
            try:
                from qrweb.models import QRRegistro
                total_qr_registros_web = QRRegistro.objects.count()
                QRRegistro.objects.all().delete()
                total_eliminados = total_qr_registros + total_qr_registros_web
            except ImportError:
                total_eliminados = total_qr_registros
            
            print(f"🗑️ ELIMINACIÓN MASIVA: {total_eliminados} activos eliminados por el usuario")
            
            return JsonResponse({
                'success': True,
                'message': f'Se eliminaron {total_eliminados} activos correctamente',
                'total_eliminados': total_eliminados
            })
            
        except Exception as e:
            print(f"❌ Error al eliminar todos los activos: {e}")
            return JsonResponse({
                'success': False,
                'message': f'Error al eliminar activos: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
